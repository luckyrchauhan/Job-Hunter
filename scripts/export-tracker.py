#!/usr/bin/env python3
"""
M6 — Export Tracker
Reads:  data/applications.json
Writes: outputs/tracker.xlsx

Columns: Company, Role, Score, Visa, Status, Applied Date, Follow-up Due,
         Urgency, Salary, Outreach Contact, Outreach Sent, Apply URL, Notes, Outcome

Usage:
  python scripts/export-tracker.py
  python scripts/export-tracker.py --output outputs/my-tracker.xlsx
"""

import json
import sys
import argparse
from datetime import datetime, timezone, date
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("✗ openpyxl not installed — run: pip install openpyxl")
    sys.exit(1)

BASE_DIR = Path(__file__).parent.parent
APPS_FILE   = BASE_DIR / "data" / "applications.json"
TRACKER_OUT = BASE_DIR / "outputs" / "tracker.xlsx"
TODAY       = date.today().isoformat()

# ─── Color palette ────────────────────────────────────────────────────────────

COLORS = {
    "header_bg":    "1F4E79",   # dark blue
    "header_fg":    "FFFFFF",
    "applied":      "D6E4F0",   # light blue
    "interviewing": "D5F5E3",   # light green
    "offer":        "A9DFBF",   # green
    "rejected":     "FADBD8",   # light red
    "stale":        "F2F3F4",   # grey
    "withdrawn":    "FDEBD0",   # light orange
    "ready":        "FEF9E7",   # light yellow
    "overdue":      "FDEDEC",   # pink — follow-up overdue
    "row_alt":      "EBF5FB",   # alternating row
}

STATUS_COLORS = {
    "applied":        COLORS["applied"],
    "outreach_sent":  COLORS["applied"],
    "interviewing":   COLORS["interviewing"],
    "offer":          COLORS["offer"],
    "rejected":       COLORS["rejected"],
    "stale":          COLORS["stale"],
    "withdrawn":      COLORS["withdrawn"],
    "ready_to_apply": COLORS["ready"],
}

COLUMNS = [
    ("Title",            32),
    ("Company",          20),
    ("Location",         18),
    ("Posted",           12),
    ("H1B Status",       18),
    ("Score",             7),
    ("Tier",             10),
    ("Domain Hit",       14),
    ("Salary",           18),
    ("URL",              40),
    ("Slack Sent",       11),
    # Application tracking
    ("Status",           15),
    ("Applied Date",     13),
    ("Follow-up Due",    13),
    ("Urgency",          16),
    ("Outreach Contact", 22),
    ("Notes",            30),
    ("Outcome",          15),
]


def load_applications() -> list:
    if not APPS_FILE.exists():
        return []
    content = APPS_FILE.read_text().strip()
    if not content:
        return []
    return json.loads(content)


def is_overdue(follow_up_due: str | None) -> bool:
    if not follow_up_due:
        return False
    try:
        return follow_up_due < TODAY
    except Exception:
        return False


def make_workbook(apps: list) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Applications"

    # ── Header row ────────────────────────────────────────────────────────────
    header_font  = Font(bold=True, color=COLORS["header_fg"], size=11)
    header_fill  = PatternFill("solid", fgColor=COLORS["header_bg"])
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font    = header_font
        cell.fill    = header_fill
        cell.alignment = header_align
        cell.border  = border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    # ── Data rows ─────────────────────────────────────────────────────────────
    sorted_apps = sorted(
        apps,
        key=lambda x: (
            {"offer": 0, "interviewing": 1, "applied": 2, "outreach_sent": 2,
             "ready_to_apply": 3, "stale": 4, "rejected": 5, "withdrawn": 6
             }.get(x.get("status", ""), 9),
            -(x.get("score") or 0)
        )
    )

    for row_idx, app in enumerate(sorted_apps, start=2):
        status      = app.get("status", "")
        follow_due  = app.get("follow_up_due", "")
        overdue     = is_overdue(follow_due) and status in ("applied", "outreach_sent", "interviewing")
        alt_row     = row_idx % 2 == 0

        row_color = STATUS_COLORS.get(status, "FFFFFF")
        if overdue:
            row_color = COLORS["overdue"]
        elif alt_row and row_color == "FFFFFF":
            row_color = COLORS["row_alt"]

        row_fill = PatternFill("solid", fgColor=row_color)
        data_align = Alignment(vertical="center", wrap_text=False)

        values = [
            app.get("role", app.get("title", "")),
            app.get("company", ""),
            app.get("location", ""),
            app.get("posted_date", ""),
            app.get("visa_check_note", app.get("visa", "")),
            app.get("score", ""),
            app.get("score_band", ""),
            app.get("domain_hit", ""),
            app.get("salary_text", ""),
            app.get("apply_url", app.get("url", "")),
            "Yes" if app.get("slack_sent") else "No",
            status,
            app.get("applied_date", ""),
            follow_due,
            app.get("urgency", ""),
            app.get("outreach_contact", ""),
            app.get("notes", ""),
            app.get("outcome", ""),
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill      = row_fill
            cell.alignment = data_align
            cell.border    = border

        ws.row_dimensions[row_idx].height = 18

    # ── Auto-filter ───────────────────────────────────────────────────────────
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    _make_summary(ws2, apps, header_font, header_fill, border)

    return wb


def _make_summary(ws, apps, header_font, header_fill, border):
    """Stats sheet: pipeline counts + follow-up due list."""
    from collections import Counter

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 10

    # Status counts
    status_counts = Counter(a.get("status", "unknown") for a in apps)
    summary_rows = [
        ("PIPELINE SUMMARY", ""),
        ("Total Applications", len(apps)),
        ("", ""),
        ("Status", "Count"),
    ]
    for status in ["applied", "outreach_sent", "interviewing", "offer",
                   "rejected", "stale", "withdrawn", "ready_to_apply"]:
        summary_rows.append((status.replace("_", " ").title(), status_counts.get(status, 0)))

    summary_rows += [
        ("", ""),
        ("FOLLOW-UPS DUE TODAY", ""),
    ]

    due_today = [
        a for a in apps
        if is_overdue(a.get("follow_up_due"))
        and a.get("status") in ("applied", "outreach_sent", "interviewing")
    ]
    if due_today:
        for a in due_today:
            summary_rows.append((f"{a.get('company')} — {a.get('role','')[:25]}", a.get("follow_up_due","")))
    else:
        summary_rows.append(("None overdue", ""))

    for row_idx, (label, val) in enumerate(summary_rows, start=1):
        c1 = ws.cell(row=row_idx, column=1, value=label)
        c2 = ws.cell(row=row_idx, column=2, value=val)
        if label in ("PIPELINE SUMMARY", "Status", "FOLLOW-UPS DUE TODAY"):
            c1.font = header_font
            c1.fill = header_fill
            c2.font = header_font
            c2.fill = header_fill
        c1.border = border
        c2.border = border


def add_daily_tab(wb: openpyxl.Workbook, apps: list, tab_name: str):
    """
    Add (or replace) a dated snapshot tab in an existing workbook.
    Tab name format: "2026-05-25" — one tab per day, historical tabs preserved.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    header_font  = Font(bold=True, color=COLORS["header_fg"], size=11)
    header_fill  = PatternFill("solid", fgColor=COLORS["header_bg"])
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Remove existing tab with same name if present (daily refresh)
    if tab_name in wb.sheetnames:
        del wb[tab_name]

    ws = wb.create_sheet(tab_name)

    # Header
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font    = header_font
        cell.fill    = header_fill
        cell.alignment = header_align
        cell.border  = border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"

    # Data
    sorted_apps = sorted(
        apps,
        key=lambda x: (
            {"offer": 0, "interviewing": 1, "applied": 2, "outreach_sent": 2,
             "ready_to_apply": 3, "stale": 4, "rejected": 5, "withdrawn": 6
             }.get(x.get("status", ""), 9),
            -(x.get("score") or 0)
        )
    )

    data_align = Alignment(vertical="center", wrap_text=False)

    for row_idx, app in enumerate(sorted_apps, start=2):
        status     = app.get("status", "")
        follow_due = app.get("follow_up_due", "")
        overdue    = is_overdue(follow_due) and status in ("applied", "outreach_sent", "interviewing")
        alt_row    = row_idx % 2 == 0

        row_color = STATUS_COLORS.get(status, "FFFFFF")
        if overdue:
            row_color = COLORS["overdue"]
        elif alt_row and row_color == "FFFFFF":
            row_color = COLORS["row_alt"]

        row_fill = PatternFill("solid", fgColor=row_color)

        values = [
            app.get("company", ""),  app.get("role", ""),
            app.get("score", ""),    app.get("visa", ""),
            status,                  app.get("applied_date", ""),
            follow_due,              app.get("urgency", ""),
            app.get("salary_text",""), app.get("outreach_contact",""),
            "Yes" if app.get("outreach_sent") else "No",
            app.get("apply_url",""), app.get("notes",""),
            app.get("outcome",""),
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill      = row_fill
            cell.alignment = data_align
            cell.border    = border

        ws.row_dimensions[row_idx].height = 18

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"


def main():
    parser = argparse.ArgumentParser(description="Export applications.json → tracker.xlsx (daily tab)")
    parser.add_argument("--output", default=str(TRACKER_OUT), help="Output .xlsx path")
    parser.add_argument("--tab-date", default=TODAY, help="Tab name date (default: today)")
    args = parser.parse_args()

    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    apps = load_applications()

    if not apps:
        print("⚠ No applications found in data/applications.json")
        print("  Creating empty tracker...")

    # Open existing workbook or create fresh
    if out_path.exists():
        wb = openpyxl.load_workbook(out_path)
        print(f"  Updating existing tracker (existing tabs: {wb.sheetnames})")
    else:
        wb = openpyxl.Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    tab_name = args.tab_date  # e.g. "2026-05-25"

    # Always update/replace "Applications" (latest full view)
    if "Applications" in wb.sheetnames:
        del wb["Applications"]
    ws_all = wb.create_sheet("Applications", 0)  # first tab

    # Rebuild Applications tab
    header_font  = Font(bold=True, color=COLORS["header_fg"], size=11)
    header_fill  = PatternFill("solid", fgColor=COLORS["header_bg"])
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws_all.cell(row=1, column=col_idx, value=col_name)
        cell.font    = header_font
        cell.fill    = header_fill
        cell.alignment = header_align
        cell.border  = border
        ws_all.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws_all.row_dimensions[1].height = 24
    ws_all.freeze_panes = "A2"

    sorted_apps = sorted(
        apps,
        key=lambda x: (
            {"offer": 0, "interviewing": 1, "applied": 2, "outreach_sent": 2,
             "ready_to_apply": 3, "stale": 4, "rejected": 5, "withdrawn": 6
             }.get(x.get("status", ""), 9),
            -(x.get("score") or 0)
        )
    )
    data_align = Alignment(vertical="center", wrap_text=False)
    for row_idx, app in enumerate(sorted_apps, start=2):
        status     = app.get("status", "")
        follow_due = app.get("follow_up_due", "")
        overdue    = is_overdue(follow_due) and status in ("applied", "outreach_sent", "interviewing")
        row_color  = STATUS_COLORS.get(status, "FFFFFF")
        if overdue:
            row_color = COLORS["overdue"]
        elif row_idx % 2 == 0 and row_color == "FFFFFF":
            row_color = COLORS["row_alt"]
        row_fill = PatternFill("solid", fgColor=row_color)
        values = [
            app.get("company",""), app.get("role",""), app.get("score",""),
            app.get("visa",""), status, app.get("applied_date",""), follow_due,
            app.get("urgency",""), app.get("salary_text",""),
            app.get("outreach_contact",""),
            "Yes" if app.get("outreach_sent") else "No",
            app.get("apply_url",""), app.get("notes",""), app.get("outcome",""),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws_all.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = row_fill; cell.alignment = data_align; cell.border = border
        ws_all.row_dimensions[row_idx].height = 18
    ws_all.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # Add dated snapshot tab
    add_daily_tab(wb, apps, tab_name)

    # Update Summary tab
    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    ws_sum = wb.create_sheet("Summary")
    _make_summary(ws_sum, apps, header_font, header_fill, border)

    wb.save(out_path)

    print(f"✅ Tracker exported → {out_path.relative_to(BASE_DIR)}")
    print(f"   Tab '{tab_name}' added/updated | Total tabs: {wb.sheetnames}")
    print(f"   {len(apps)} application(s)")

    from collections import Counter
    counts = Counter(a.get("status","") for a in apps)
    if counts:
        for status, n in sorted(counts.items()):
            print(f"   {status:<20} {n}")


if __name__ == "__main__":
    main()
